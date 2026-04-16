"""
Build ОКПД2 coverage map from РРПП (ПП 719) and РЭП (ПП 878) registry exports.

Usage:
    python scripts/build_registry_coverage.py \
        --pp719 "path/to/production.xlsx" \
        --pp878 "path/to/production-rep.xlsx" \
        --out app/data/pp_registries_coverage.json

What the script produces (per-ОКПД2):
  - which registries a code shows up in (pp719 / pp878 / both)
  - which Минпромторг departments issue conclusions for it
  - record counts in each registry
  - score statistics from the "Баллы" column (pp719 — actual scores, not thresholds)
  - level distribution from the "Уровень радиоэлектронной продукции" column (pp878)

Это НЕ справочник минимальных порогов. Это карта того, что реально лежит
в реестрах — используем как baseline для валидации будущего парсера ПП 719/878/1875.
"""
from __future__ import annotations

import argparse
import json
import statistics
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Optional

import openpyxl


# Column positions are 1-indexed in openpyxl; we store 0-indexed Python tuples.
# Both registries put headers on row 3 (row 1 = export timestamp banner, row 2 = blank).
HEADER_ROW = 3
DATA_START_ROW = 4


def _iter_data_rows(path: Path) -> Iterable[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            yield row
    finally:
        wb.close()


def _read_header(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        for row in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True):
            return [str(v).strip() if v is not None else "" for v in row]
    finally:
        wb.close()
    return []


def _resolve_columns(header: list[str], wanted: dict[str, list[str]]) -> dict[str, Optional[int]]:
    """
    Map semantic names to 0-indexed column positions by looking up candidate
    header labels. Registries have slightly different headers (e.g. 'ОКПД 2'
    vs 'ОКПД2'), so we try each candidate until one matches.
    """
    norm = [h.lower().replace(" ", "").replace("\xa0", "") for h in header]
    resolved: dict[str, Optional[int]] = {}
    for key, candidates in wanted.items():
        idx: Optional[int] = None
        for cand in candidates:
            cnorm = cand.lower().replace(" ", "").replace("\xa0", "")
            try:
                idx = norm.index(cnorm)
                break
            except ValueError:
                continue
        resolved[key] = idx
    return resolved


def _cell(row: tuple, idx: Optional[int]):
    """Safe row[idx] — tolerates short rows (openpyxl trims trailing None)."""
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_score(raw) -> Optional[float]:
    if raw is None or raw == "":
        return None
    try:
        return float(str(raw).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return None


def _normalize_okpd2(raw) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # Some registries pad with non-breaking spaces or extra dots.
    s = s.replace("\xa0", " ").strip(" .")
    return s or None


def _summarize_scores(scores: list[float]) -> dict:
    if not scores:
        return {"count": 0}
    scores_sorted = sorted(scores)
    n = len(scores_sorted)
    return {
        "count": n,
        "min": scores_sorted[0],
        "max": scores_sorted[-1],
        "median": statistics.median(scores_sorted),
        "p25": scores_sorted[n // 4],
        "p75": scores_sorted[(3 * n) // 4],
    }


def process_pp719(path: Path, coverage: dict) -> dict:
    header = _read_header(path)
    cols = _resolve_columns(
        header,
        {
            "okpd2": ["ОКПД2", "ОКПД 2"],
            "score": ["Баллы"],
            "percent": ["Процентный показатель"],
            "department": ["Заключение: Департамент"],
            "registry_number": ["Реестровый номер"],
        },
    )
    if cols["okpd2"] is None:
        raise RuntimeError(f"ОКПД2 column not found in {path.name}. Headers: {header}")

    total = 0
    seen_codes: set[str] = set()
    per_code_scores: dict[str, list[float]] = defaultdict(list)
    per_code_departments: dict[str, set[str]] = defaultdict(set)

    for row in _iter_data_rows(path):
        okpd2 = _normalize_okpd2(_cell(row, cols["okpd2"]))
        if not okpd2:
            continue
        total += 1
        seen_codes.add(okpd2)

        entry = coverage.setdefault(okpd2, {"pp719_records": 0, "pp878_records": 0})
        entry["pp719_records"] += 1

        score = _parse_score(_cell(row, cols["score"]))
        if score is not None:
            per_code_scores[okpd2].append(score)

        dept = _cell(row, cols["department"])
        if dept:
            per_code_departments[okpd2].add(str(dept).strip())

    for code, scores in per_code_scores.items():
        coverage[code]["pp719_score_stats"] = _summarize_scores(scores)
    for code, depts in per_code_departments.items():
        coverage[code].setdefault("departments", set()).update(depts)

    return {
        "file": str(path),
        "rows_processed": total,
        "unique_okpd2": len(seen_codes),
    }


def process_pp878(path: Path, coverage: dict) -> dict:
    header = _read_header(path)
    cols = _resolve_columns(
        header,
        {
            "okpd2": ["ОКПД2", "ОКПД 2"],
            "score": ["Баллы"],
            "percent": ["Процентный показатель"],
            "level": ["Уровень радиоэлектронной продукции"],
            "department": ["Заключение: Департамент"],
        },
    )
    if cols["okpd2"] is None:
        raise RuntimeError(f"ОКПД2 column not found in {path.name}. Headers: {header}")

    total = 0
    seen_codes: set[str] = set()
    per_code_scores: dict[str, list[float]] = defaultdict(list)
    per_code_levels: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    per_code_departments: dict[str, set[str]] = defaultdict(set)

    for row in _iter_data_rows(path):
        okpd2 = _normalize_okpd2(_cell(row, cols["okpd2"]))
        if not okpd2:
            continue
        total += 1
        seen_codes.add(okpd2)

        entry = coverage.setdefault(okpd2, {"pp719_records": 0, "pp878_records": 0})
        entry["pp878_records"] += 1

        score = _parse_score(_cell(row, cols["score"]))
        if score is not None:
            per_code_scores[okpd2].append(score)

        level = _cell(row, cols["level"])
        if level:
            per_code_levels[okpd2][str(level).strip()] += 1

        dept = _cell(row, cols["department"])
        if dept:
            per_code_departments[okpd2].add(str(dept).strip())

    for code, scores in per_code_scores.items():
        coverage[code]["pp878_score_stats"] = _summarize_scores(scores)
    for code, levels in per_code_levels.items():
        coverage[code]["pp878_level_distribution"] = dict(levels)
    for code, depts in per_code_departments.items():
        coverage[code].setdefault("departments", set()).update(depts)

    return {
        "file": str(path),
        "rows_processed": total,
        "unique_okpd2": len(seen_codes),
    }


def finalize(coverage: dict) -> dict:
    """Convert sets to sorted lists and attach 'registries' summary field."""
    for code, entry in coverage.items():
        registries = []
        if entry.get("pp719_records", 0) > 0:
            registries.append("pp719")
        if entry.get("pp878_records", 0) > 0:
            registries.append("pp878")
        entry["registries"] = registries
        if "departments" in entry:
            entry["departments"] = sorted(entry["departments"])
    return coverage


def build_summary(coverage: dict, pp719_info: dict, pp878_info: dict) -> dict:
    only_719 = sum(1 for e in coverage.values() if e["registries"] == ["pp719"])
    only_878 = sum(1 for e in coverage.values() if e["registries"] == ["pp878"])
    both = sum(1 for e in coverage.values() if set(e["registries"]) == {"pp719", "pp878"})

    dept_counter: dict[str, int] = defaultdict(int)
    for e in coverage.values():
        for d in e.get("departments", []):
            dept_counter[d] += 1
    top_departments = sorted(dept_counter.items(), key=lambda kv: -kv[1])[:15]

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "pp719": pp719_info,
        "pp878": pp878_info,
        "unique_okpd2_total": len(coverage),
        "okpd2_only_in_pp719": only_719,
        "okpd2_only_in_pp878": only_878,
        "okpd2_in_both": both,
        "top_departments_by_okpd2_coverage": [
            {"department": d, "okpd2_count": c} for d, c in top_departments
        ],
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--pp719", type=Path, required=True, help="Path to РРПП export (production.xlsx)")
    parser.add_argument("--pp878", type=Path, required=True, help="Path to РЭП export (production-rep.xlsx)")
    parser.add_argument("--out", type=Path, required=True, help="Output JSON path")
    args = parser.parse_args()

    for p in (args.pp719, args.pp878):
        if not p.exists():
            print(f"ERROR: input file not found: {p}", file=sys.stderr)
            return 2

    coverage: dict = {}

    print(f"[pp719] processing {args.pp719.name} ...", flush=True)
    pp719_info = process_pp719(args.pp719, coverage)
    print(f"[pp719] rows={pp719_info['rows_processed']:,} unique_okpd2={pp719_info['unique_okpd2']:,}")

    print(f"[pp878] processing {args.pp878.name} ...", flush=True)
    pp878_info = process_pp878(args.pp878, coverage)
    print(f"[pp878] rows={pp878_info['rows_processed']:,} unique_okpd2={pp878_info['unique_okpd2']:,}")

    finalize(coverage)
    summary = build_summary(coverage, pp719_info, pp878_info)

    output = {"_summary": summary, "okpd2": dict(sorted(coverage.items()))}
    args.out.parent.mkdir(parents=True, exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print()
    print("=== SUMMARY ===")
    print(f"  unique ОКПД2 total: {summary['unique_okpd2_total']:,}")
    print(f"  only in ПП 719:     {summary['okpd2_only_in_pp719']:,}")
    print(f"  only in ПП 878:     {summary['okpd2_only_in_pp878']:,}")
    print(f"  in both registries: {summary['okpd2_in_both']:,}")
    print()
    print("  Top-15 departments by ОКПД2 coverage:")
    for d in summary["top_departments_by_okpd2_coverage"]:
        print(f"    {d['okpd2_count']:>5}  {d['department']}")
    print()
    print(f"  Written: {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
